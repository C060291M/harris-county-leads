import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import os

conn = psycopg2.connect(os.environ["DATABASE_URL"], connect_timeout=30)
cur = conn.cursor()

wb = openpyxl.load_workbook("bell_export.xlsx", read_only=True)
ws = wb[wb.sheetnames[0]]

rows_iter = ws.iter_rows(values_only=True)
header = next(rows_iter)
idx = {name: i for i, name in enumerate(header)}

batch = []
total = 0
for row in rows_iter:
    if row[idx["prop_type_cd"]] != "R":
        continue
    batch.append((
        str(row[idx["prop_id"]]) if row[idx["prop_id"]] else None,
        row[idx["Owner_name"]],
        row[idx["Ownr_Addr"]],
        row[idx["Situs_Addr"]],
        str(row[idx["living_area"]]) if row[idx["living_area"]] else None,
        str(row[idx["yr_blt"]]) if row[idx["yr_blt"]] else None,
        str(row[idx["market"]]) if row[idx["market"]] else None,
        row[idx["legal_desc"]],
        row[idx["subdiv"]],
    ))
    if len(batch) >= 2000:
        execute_values(cur, "INSERT INTO bell_cad (prop_id, owner_name, owner_address, situs_address, living_area, yr_built, market_value, legal_desc, subdiv) VALUES %s", batch)
        conn.commit()
        total += len(batch)
        print(f"{total} rows imported so far...")
        batch = []

if batch:
    execute_values(cur, "INSERT INTO bell_cad (prop_id, owner_name, owner_address, situs_address, living_area, yr_built, market_value, legal_desc, subdiv) VALUES %s", batch)
    conn.commit()
    total += len(batch)

print(f"DONE: {total} residential records imported")
